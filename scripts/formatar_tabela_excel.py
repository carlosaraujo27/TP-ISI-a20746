import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

# Caminho do ficheiro Excel que foi gerado pelo Excel Writer
file_path = 'C:/Users/Carlos/knime-workspace/TP-ISI-a20746/scripts/eventos_futuros.xlsx'

# Carregar o ficheiro Excel
book = openpyxl.load_workbook(file_path)
sheet = book.active  # Usar a primeira folha ativa ou especificar o nome da folha, se necessário

# Remover a tabela existente com o nome "EventosFuturos" (caso exista)
existing_tables = [tbl for tbl in sheet._tables if hasattr(tbl, 'name') and tbl.name == "EventosFuturos"]
for tbl in existing_tables:
    sheet._tables.remove(tbl)

# Determinar o intervalo da nova tabela (exemplo: A1 até a última célula com dados)
min_col, max_col = 1, sheet.max_column
min_row, max_row = 1, sheet.max_row
table_range = f"A{min_row}:{openpyxl.utils.get_column_letter(max_col)}{max_row}"

# Criar a nova tabela no intervalo de dados
table = Table(displayName="EventosFuturos", ref=table_range)

# Estilizar a tabela
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, 
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
table.tableStyleInfo = style

# Adicionar a nova tabela à folha
sheet.add_table(table)

# Autosize das colunas e centralização das células
for col in sheet.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Obter a letra da coluna
    for cell in col:
        # Centralizar o conteúdo de cada célula
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Calcular o tamanho máximo do conteúdo da coluna
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    # Ajustar a largura da coluna com base no tamanho máximo
    adjusted_width = (max_length + 2)  # Ajuste adicional para melhor visualização
    sheet.column_dimensions[col_letter].width = adjusted_width

# Salvar o ficheiro formatado
book.save(file_path)
